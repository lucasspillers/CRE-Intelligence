from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO


def build_excel_model(
    property_name,
    purchase_price,
    annual_noi,
    loan_to_value,
    interest_rate,
    hold_period,
    exit_cap_rate,
    noi_growth_rate,
    amortization_period,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Model"

    input_font = Font(color="0000FF")
    label_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws["A1"] = "Property Name"
    ws["B1"] = property_name
    ws["A1"].font = label_font

    ws["A2"] = "Purchase Price"
    ws["B2"] = purchase_price
    ws["B2"].font = input_font
    ws["B2"].number_format = '$#,##0'

    ws["A3"] = "Annual NOI (Year 1)"
    ws["B3"] = annual_noi
    ws["B3"].font = input_font
    ws["B3"].number_format = '$#,##0'

    ws["A4"] = "Loan-to-Value"
    ws["B4"] = loan_to_value / 100
    ws["B4"].font = input_font
    ws["B4"].number_format = '0.0%'

    ws["A5"] = "Interest Rate"
    ws["B5"] = interest_rate / 100
    ws["B5"].font = input_font
    ws["B5"].number_format = '0.0%'

    ws["A6"] = "Hold Period (years)"
    ws["B6"] = hold_period
    ws["B6"].font = input_font

    ws["A7"] = "Exit Cap Rate"
    ws["B7"] = exit_cap_rate / 100
    ws["B7"].font = input_font
    ws["B7"].number_format = '0.0%'

    ws["A8"] = "NOI Growth Rate"
    ws["B8"] = noi_growth_rate / 100
    ws["B8"].font = input_font
    ws["B8"].number_format = '0.0%'

    ws["A9"] = "Loan Amortization Period (years)"
    ws["B9"] = amortization_period
    ws["B9"].font = input_font

    ws["A11"] = "Loan Amount"
    ws["B11"] = "=B2*B4"
    ws["B11"].number_format = '$#,##0'

    ws["A12"] = "Initial Equity"
    ws["B12"] = "=B2-B11"
    ws["B12"].number_format = '$#,##0'

    ws["A13"] = "Annual Debt Service"
    ws["B13"] = "=B11*B5/(1-(1+B5)^(-B9))"
    ws["B13"].number_format = '$#,##0'

    ws["A15"] = "Year"
    ws["B15"] = "NOI"
    ws["C15"] = "Debt Service"
    ws["D15"] = "Cash Flow"
    for cell in ["A15", "B15", "C15", "D15"]:
        ws[cell].font = label_font
        ws[cell].fill = header_fill

    start_row = 16
    for i in range(int(hold_period)):
        row = start_row + i
        ws[f"A{row}"] = i + 1
        if i == 0:
            ws[f"B{row}"] = "=$B$3"
        else:
            ws[f"B{row}"] = f"=B{row - 1}*(1+$B$8)"
        ws[f"B{row}"].number_format = '$#,##0'
        ws[f"C{row}"] = "=$B$13"
        ws[f"C{row}"].number_format = '$#,##0'
        ws[f"D{row}"] = f"=B{row}-C{row}"
        ws[f"D{row}"].number_format = '$#,##0'

    end_row = start_row + int(hold_period) - 1

    summary_row = end_row + 2
    ws[f"A{summary_row}"] = "Total Cash Flow Over Hold"
    ws[f"B{summary_row}"] = f"=SUM(D{start_row}:D{end_row})"
    ws[f"B{summary_row}"].number_format = '$#,##0'

    ws[f"A{summary_row + 1}"] = "Projected Exit NOI"
    ws[f"B{summary_row + 1}"] = f"=B{end_row}*(1+$B$8)"
    ws[f"B{summary_row + 1}"].number_format = '$#,##0'

    ws[f"A{summary_row + 2}"] = "Exit Value"
    ws[f"B{summary_row + 2}"] = f"=B{summary_row + 1}/$B$7"
    ws[f"B{summary_row + 2}"].number_format = '$#,##0'
    
    ws[f"A{summary_row + 3}"] = "Remaining Loan Balance"
    ws[f"B{summary_row + 3}"] = f"=B11*(1+B5)^B6-B13*(((1+B5)^B6-1)/B5)"
    ws[f"B{summary_row + 3}"].number_format = '$#,##0'

    ws[f"A{summary_row + 4}"] = "Net Sale Proceeds"
    ws[f"B{summary_row + 4}"] = f"=B{summary_row + 2}-B{summary_row + 3}"
    ws[f"B{summary_row + 4}"].number_format = '$#,##0'

    ws[f"A{summary_row + 5}"] = "Total Equity Value"
    ws[f"B{summary_row + 5}"] = f"=B12+B{summary_row}+B{summary_row + 4}"
    ws[f"B{summary_row + 5}"].number_format = '$#,##0'

    ws[f"A{summary_row + 6}"] = "Equity Multiple"
    ws[f"B{summary_row + 6}"] = f"=B{summary_row + 5}/B12"
    ws[f"B{summary_row + 6}"].number_format = '0.00"x"'

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
